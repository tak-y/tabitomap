import sys, os, os.path
from flask import *
from werkzeug.utils import secure_filename
import pandas as pd
import folium
import shutil
from folium import plugins
from livereload import Server
import googlemaps
import branca
import math
import openpyxl
import geocoder
import sys, os
sys.path.insert(0, '/var/www/html')

app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
allowed_ext = {'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in request.files['file'] and filename.rsplit(',',1).lower() in allowed_ext
@app.route('/')
def index():
    return render_template('index.html')
@app.route('/visual', methods=['GET','POST'])
def visual():
    if request.method == 'POST':
        googleapikey = 'AIzaSyC8rTI8Yv1LrEDnRJ109DfpUsdaQBXAfhE'
        gmaps = googlemaps.Client(key=googleapikey)
        file = request.files['file']
        print(file)
        filename = secure_filename(file.filename)
        file.save(os.path.join('upload', filename))
        if filename.split('.')[1] in allowed_ext:

            flash('file upload')
            wb=openpyxl.load_workbook(file)
            ws = wb["Sheet1"]
            popup = str(request.form['name'])
            link = str(request.form['data'])
            num = math.floor((ws.max_row-1)/2)+1
            print('saidai'+str(ws.max_row))

            m = folium.Map(location=[geocoder.osm(ws.cell(row=2,column=1).value,timeout=1.0).latlng[0],geocoder.osm(ws.cell(row=2,column=1).value,timeout=1.0).latlng[1]],zoom_start=13) 
            for row in ws.iter_rows(min_row=2,max_col=2,max_row=ws.max_row):
                for cell in row:
                    if cell.row<=ws.max_row:
                        orig0 = gmaps.geocode(ws.cell(row=cell.row,column=1).value)
                        print(orig0[0])
                        try:
                            o_O_lat = orig0[0].get('geometry').get('location').get('lat')
                            o_O_lng = orig0[0].get('geometry').get('location').get('lng')
                            print('try1')
                        except AttributeError:
                            try:
                                o_O_lat = orig0[0]['geometry']['location']['lat']
                                o_O_lng = orig0[0]['geometry']['location']['lng']
                                print('try2')
                            except KeyError:
                                o_O_lat = orig0['geometry']['location']['lat']
                                o_O_lng = orig0['geometry']['location']['lng']
                                print('try3')
                        except KeyError:
                            try:
                                o_O_lat = orig0['geometry']['location']['lat']
                                o_O_lng = orig0['geometry']['location']['lng']
                                print('try4')
                            except AttributeError:
                                o_O_lat = orig0[0]['geometry']['location']['lat']
                                o_O_lng = orig0[0]['geometry']['location']['lng']
                        url=str(ws.cell(row=cell.row,column=2).value)
                        print(url)
                        pop=str(ws.cell(row=cell.row,column=1).value)

                        html='<a href ="{url}" target="_blank" rel="noopener noreferrer">{pop}</a>'.format(url=url,pop=pop)
                        iframe = branca.element.IFrame(html=html, width=300, height=100)
                        popup = folium.Popup(iframe, max_width=300,show=True)
                            
                        folium.Marker(location=[o_O_lat,o_O_lng],color="blue",popup=popup).add_to(m)
                        m.save('static/data_r.html')
                    else:
                        break


            return render_template('visual.html')

        else:
            print(error)

if __name__ == "__main__":
    app.run(debug=True, port=8000, host='0.0.0.0') 





